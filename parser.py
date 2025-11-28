import openpyxl
import re
from typing import Dict, List, Any


def load_ws(file):
    """
    从 Streamlit 上传的 xlsx 文件里读取第一个工作表。
    file 可以是路径，也可以是 UploadedFile 对象。
    """
    wb = openpyxl.load_workbook(file, data_only=True)
    return wb[wb.sheetnames[0]]


def get_month_columns(ws) -> Dict[str, int]:
    """
    读取第 1 行，从第 6 列开始，所有包含“月”的单元格，构造成：
    { '2024年11月份': 6, '2024年12月份': 7, '25年3月份': 8, ... }
    """
    header_row = 1
    month_map: Dict[str, int] = {}

    for col in range(6, 50):  # F 列开始向右扫
        val = ws.cell(header_row, col).value
        if not isinstance(val, str):
            continue
        text = val.strip()
        if not text:
            continue
        if "月" in text:
            month_map[text] = col

    return month_map


def split_cell_into_people(text: str) -> List[str]:
    """
    把一个单元格里的内容按“多人推荐”拆分成多个片段，每个片段代表一个人。

    规则：
    - 先按行拆开，去掉空行。
    - 识别“头行”（认为是一个人的起始行）：
        * 含“推荐”的行；
        * 或者形如 “XX-XX之星” 的行（不写“推荐”也能识别）。
    - 从每个头行开始，直到下一个头行之前的所有行，视为一个人的完整文本片段。
    """
    if not isinstance(text, str):
        text = str(text or "")

    lines = [l.strip() for l in text.splitlines() if l and str(l).strip()]
    if not lines:
        return []

    header_idx: List[int] = []

    for i, line in enumerate(lines):
        # 情况1：包含“推荐”
        if "推荐" in line:
            header_idx.append(i)
            continue

        # 情况2：不含“推荐”，但长得像 “姓名-xxx之星”
        #   例： "卢俊宏-敬业之星"
        if "之星" in line and re.search(r'[\u4e00-\u9fff]{2,4}\s*[-－:：]', line):
            header_idx.append(i)

    header_idx = sorted(set(header_idx))

    # 如果没识别出头行，就当成一个整体
    if not header_idx:
        return [text.strip()]

    segments: List[str] = []
    for j, start in enumerate(header_idx):
        end = header_idx[j + 1] if j + 1 < len(header_idx) else len(lines)
        seg_lines = lines[start:end]
        seg = "\n".join(seg_lines).strip()
        if seg:
            segments.append(seg)

    return segments


def parse_name_award(text: str):
    """
    从一整段“推荐 + 奖项”文本中尽量拆出：
    - name: 姓名
    - award: 奖项名称（XXX之星）

    兼容格式：
    - '推荐：张三-突出贡献之星'
    - '推荐张三：敬业之星'
    - '推荐张三-敬业之星\\n评语：……'
    - '张三-敬业之星'
    - '张三：敬业之星'
    """
    t = (text or "").strip()
    if not t:
        return "", ""

    first_line = t.splitlines()[0].strip()

    # 去掉开头的“推荐”
    for prefix in ["推荐：", "推荐:", "推荐 ", "推荐"]:
        if first_line.startswith(prefix):
            first_line = first_line[len(prefix):].strip()
            break

    # 常见分隔符：张三-敬业之星 / 张三：敬业之星
    for sep in ["：", ":", "-", "－"]:
        if sep in first_line:
            name, award = first_line.split(sep, 1)
            return name.strip("【】 、， "), award.strip()

    # 兜底：取前 2~3 字作姓名，后面当奖项
    plain = first_line.strip("【】 、， ")
    if len(plain) >= 3:
        return plain[:2], plain[2:]
    return plain, ""


def parse_comment(text: str) -> str:
    """
    抽取评语：
    - 如果包含“评语”，取“评语”后面的内容
    - 否则，如果有多行，取第 2 行开始
    - 再否则，就返回整段文本
    """
    t = (text or "").strip()
    if not t:
        return ""

    if "评语" in t:
        idx = t.find("评语")
        sub = t[idx + len("评语") :]
        sub = sub.lstrip("：:").strip()
        return sub

    lines = [l.strip() for l in t.splitlines() if l.strip()]
    if len(lines) > 1:
        return "\n".join(lines[1:])
    return t


def extract(ws, col: int) -> List[Dict[str, Any]]:
    """
    从指定月份列（col）里抽取所有“每月之星”记录。

    约定：
    - 行 3 开始是数据（A 列序号为 1,2,3…）
    - A 列为空视为数据结束
    - 对应月份列为空 or “本次暂无” → 跳过
    - 如果一个格子里推荐了多个人（多段“推荐XX-XX之星”），拆成多条记录。
    """
    results: List[Dict[str, Any]] = []
    row = 3

    while True:
        seq = ws.cell(row, 1).value
        if seq is None:
            break  # 序号为空，认为到尾部了

        raw = ws.cell(row, col).value
        if raw is None:
            row += 1
            continue

        text = str(raw).strip()
        if (not text) or text == "本次暂无":
            row += 1
            continue

        dept1 = ws.cell(row, 2).value or ""
        dept2 = ws.cell(row, 3).value or ""

        # 🔥 关键：这里拆多人
        segments = split_cell_into_people(text)
        if not segments:
            row += 1
            continue

        for seg in segments:
            name, award = parse_name_award(seg)
            comment = parse_comment(seg)

            # 垃圾段落过滤一下：没有姓名就丢弃
            if not name or name in ("推荐", "评语"):
                continue

            results.append(
                {
                    "row": row,
                    "dept1": str(dept1),
                    "dept2": str(dept2),
                    "name": name,
                    "award": award,
                    "comment": comment,
                    "raw": seg,  # 用拆分后的片段作为raw，更直观
                }
            )

        row += 1

    return results


