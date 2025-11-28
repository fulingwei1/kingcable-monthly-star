import openpyxl
import re
from typing import Dict, List, Any


def load_ws(file):
    """
    从上传的 Excel 文件中读取第一个工作表。
    file 可以是路径，也可以是 Streamlit 的 UploadedFile。
    """
    wb = openpyxl.load_workbook(file, data_only=True)
    return wb[wb.sheetnames[0]]


def get_month_columns(ws) -> Dict[str, int]:
    """
    读取第 1 行，从第 6 列开始，所有包含“月”的单元格，构造成：
    { '2024年11月份': 6, '2024年12月份': 7, '25年10月份': 8, ... }
    """
    header_row = 1
    month_map: Dict[str, int] = {}

    # F 列开始往右扫一大段，够用
    for col in range(6, 80):
        val = ws.cell(header_row, col).value
        if not isinstance(val, str):
            continue
        text = val.strip()
        if text and "月" in text:
            month_map[text] = col

    return month_map


def _split_cell_into_people(text: str) -> List[str]:
    """
    将一个单元格内容拆分成多个“个人推荐片段”。

    支持几类人头格式（可以混用）：
    - 推荐：张三-敬业之星
    - 推荐：张三-核心技术骨干
    - 张三-敬业之星 / 张三:敬业之星 / 张三：敬业之星
    - 张三【敬业之星】 / 张三【核心技术骨干】

    逻辑：在整段文本中找到所有“人头”的起始位置，再按照这些起点切分。
    """
    if not isinstance(text, str):
        text = str(text or "")
    text = text.strip()
    if not text:
        return []

    header_pattern = re.compile(
        r'(推荐[:： ]*)?'                 # 可选的“推荐”
        r'([\u4e00-\u9fff]{2,4})'        # 姓名：2~4 个汉字
        r'\s*'
        r'(?:'
        r'【[^】\n]{0,30}】'              # 张三【敬业之星】 / 张三【核心技术骨干】
        r'|[-－:：][^，。；\n]{0,30}'      # 张三-敬业之星 / 张三-核心技术骨干 / 张三：敬业之星
        r')'
    )

    matches = list(header_pattern.finditer(text))
    if not matches:
        # 完全识别不出人头，就当成一个整体，后面交给 _parse_name_award 兜底
        return [text]

    segments: List[str] = []
    for i, m in enumerate(matches):
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        seg = text[start:end].strip()
        if seg:
            segments.append(seg)

    return segments


def _parse_name_award(text: str):
    """
    从一个“个人片段”里解析姓名和奖项，返回 (name, award)。

    支持三类主格式：
    A: 推荐：张三-敬业之星
    B: 张三-敬业之星 / 张三：敬业之星
    C: 张三【敬业之星】
    """
    t = (text or "").strip()
    if not t:
        return "", ""

    first = t.splitlines()[0].strip()

    # 格式 C：张三【敬业之星】
    m = re.match(r'^([\u4e00-\u9fff]{2,4})【(.+?)】', first)
    if m:
        return m.group(1).strip(), m.group(2).strip()

    # 去掉“推荐”
    for prefix in ["推荐：", "推荐:", "推荐 ", "推荐"]:
        if first.startswith(prefix):
            first = first[len(prefix):].strip()
            break

    # 格式 A & B：张三-敬业之星 / 张三：敬业之星
    for sep in ["：", ":", "-", "－"]:
        if sep in first:
            name, award = first.split(sep, 1)
            return name.strip(), award.strip()

    # 兜底：不符合以上格式，就按前 2 字做姓名，剩下做奖项
    first = first.strip("【】 、， ")
    if len(first) >= 3:
        return first[:2], first[2:]
    return first, ""


def _parse_comment(text: str) -> str:
    """
    抽取评语：
    - 如果包含“评语”，取“评语”后面的内容
    - 否则，如果有多行，取第 2 行开始
    - 再否则，就返回整段文本
    """
    t = (text or "").strip()
    if not t:
        return ""

    if "评语" in t:
        idx = t.find("评语")
        sub = t[idx + len("评语") :]
        sub = sub.lstrip("：:").strip()
        return sub

    lines = [l.strip() for l in t.splitlines() if l.strip()]
    if len(lines) > 1:
        return "\n".join(lines[1:])
    return t


def extract(ws, col: int) -> List[Dict[str, Any]]:
    """
    从指定月份列（col）里抽取所有“每月之星”记录。

    约定：
    - 行 3 开始是数据（A 列序号为 1,2,3…）
    - A 列为空视为数据结束
    - 对应月份列为空 or “本次暂无” → 跳过
    - 如果一个格子里推荐了多个人，拆成多条记录。
    """
    results: List[Dict[str, Any]] = []
    row = 3

    while True:
        seq = ws.cell(row, 1).value
        if seq is None:
            break  # 序号为空，认为到尾部了

        raw = ws.cell(row, col).value
        if raw is None:
            row += 1
            continue

        text = str(raw).strip()
        if (not text) or text == "本次暂无":
            row += 1
            continue

        dept1 = ws.cell(row, 2).value or ""
        dept2 = ws.cell(row, 3).value or ""

        # 拆出多人
        segments = _split_cell_into_people(text)
        if not segments:
            row += 1
            continue

        for seg in segments:
            name, award = _parse_name_award(seg)
            comment = _parse_comment(seg)

            if not name or name in ("推荐", "评语"):
                continue

            results.append(
                {
                    "row": row,
                    "dept1": str(dept1),
                    "dept2": str(dept2),
                    "name": name,
                    "award": award,
                    "comment": comment,
                    "raw": seg,
                }
            )

        row += 1

    return results
